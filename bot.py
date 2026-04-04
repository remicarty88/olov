import logging
import asyncio
import re
import os
from datetime import datetime
import pandas as pd
import aiohttp
from typing import Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, FSInputFile
from aiogram.dispatcher.middlewares.base import BaseMiddleware

# --- CONFIGURATION ---
BOT_TOKEN = "8752266798:AAFlBiA2F9_xlmymr6yFK0ENzZ9fHiwZutA"
FIREBASE_URL = "https://neonapp-a05b0-default-rtdb.firebaseio.com/"

ACCESS_PIN = "1188"
ADMIN_ID = 6201234513
WIPE_PIN = "88"

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- FIREBASE (Realtime Database) REST CLIENT ---
# ПРИМЕЧАНИЕ: Работа без serviceAccountKey.json возможна ТОЛЬКО если в Firebase Realtime Database
# установлены публичные правила (Rules) на время тестирования.
# Для продакшена нужно включать Auth и/или использовать Admin SDK с serviceAccountKey.json.


class FirebaseREST:
    def __init__(self, base_url: str, session: aiohttp.ClientSession):
        self._base_url = base_url.rstrip("/")
        self._session = session

    def _url(self, path: str) -> str:
        path = path.strip("/")
        return f"{self._base_url}/{path}.json"

    async def get(self, path: str):
        async with self._session.get(self._url(path)) as resp:
            resp.raise_for_status()
            return await resp.json()

    async def put(self, path: str, payload: dict):
        async with self._session.put(self._url(path), json=payload) as resp:
            resp.raise_for_status()
            return await resp.json()

    async def patch(self, path: str, payload: dict):
        async with self._session.patch(self._url(path), json=payload) as resp:
            resp.raise_for_status()
            return await resp.json()

    async def post(self, path: str, payload: dict):
        async with self._session.post(self._url(path), json=payload) as resp:
            resp.raise_for_status()
            return await resp.json()

    async def delete(self, path: str):
        async with self._session.delete(self._url(path)) as resp:
            resp.raise_for_status()
            return await resp.json()


firebase: Optional[FirebaseREST] = None
firebase_session: Optional[aiohttp.ClientSession] = None

_auth_cache: dict[int, bool] = {}

# Регистрация шрифта Arial для поддержки кириллицы в PDF
FONT_PATH = "/System/Library/Fonts/Supplemental/Arial.ttf"
if os.path.exists(FONT_PATH):
    pdfmetrics.registerFont(TTFont('Arial', FONT_PATH))
    DEFAULT_FONT = 'Arial'
else:
    DEFAULT_FONT = 'Helvetica' # Fallback

# --- FSM STATES ---
class Form(StatesGroup):
    add_supplier = State()
    select_supplier_for_debt = State()
    amount_debt = State()
    select_supplier_for_payment = State()
    amount_payment = State()
    auth_pin = State()
    admin_wipe_pin = State()

# --- UTILS ---
def format_currency(amount):
    return f"{amount:,.0f}".replace(",", " ") + " UZS"

def clean_amount(amount_str):
    # Remove everything except digits
    cleaned = re.sub(r'[^\d]', '', amount_str)
    return int(cleaned) if cleaned else 0


def _normalize_cmd(text: str) -> str:
    if not text:
        return ""
    t = text.strip().lower()
    t = re.sub(r"[\s\t\n]+", " ", t)
    t = re.sub(r"^[^a-zа-я0-9]+\s*", "", t)
    return t


def cmd_filter(*variants: str):
    normalized = {_normalize_cmd(v) for v in variants}
    return lambda m: _normalize_cmd(getattr(m, "text", "")) in normalized


def is_admin(user_id: int) -> bool:
    return user_id == ADMIN_ID


async def is_authorized(user_id: int) -> bool:
    if user_id in _auth_cache:
        return _auth_cache[user_id]
    if firebase is None:
        return False
    try:
        data = await firebase.get(f"auth_users/{user_id}")
        ok = bool(data and isinstance(data, dict) and data.get("authorized") is True)
        _auth_cache[user_id] = ok
        return ok
    except Exception:
        _auth_cache[user_id] = False
        return False


async def set_authorized(user_id: int, value: bool):
    if firebase is None:
        return
    await firebase.put(f"auth_users/{user_id}", {
        "authorized": bool(value),
        "updated_at": datetime.now().isoformat(),
    })
    _auth_cache[user_id] = bool(value)


class AuthMiddleware(BaseMiddleware):
    async def __call__(self, handler, event, data):
        if not isinstance(event, types.Message):
            return await handler(event, data)

        message: types.Message = event
        user_id = message.from_user.id if message.from_user else 0
        state: FSMContext = data.get("state")

        text = message.text or ""
        is_start = text.startswith("/start")
        current_state = await state.get_state() if state else None

        if is_start:
            return await handler(event, data)

        # allow entering pins while in pin states
        if current_state in {Form.auth_pin.state, Form.admin_wipe_pin.state}:
            return await handler(event, data)

        # All other interactions require authorization
        if not await is_authorized(user_id):
            if state:
                await state.set_state(Form.auth_pin)
            await message.answer("🔐 Введите PIN-код для доступа:", reply_markup=ReplyKeyboardRemove())
            return

        return await handler(event, data)


async def build_detailed_report():
    if firebase is None:
        raise RuntimeError("Firebase not initialized")

    suppliers = await firebase.get('suppliers') or {}
    history = await firebase.get('history') or {}

    agg = {}
    for _, item in history.items():
        if not isinstance(item, dict):
            continue
        supplier_id = item.get('supplier_id')
        if not supplier_id:
            continue
        amount = int(item.get('amount', 0) or 0)
        if supplier_id not in agg:
            agg[supplier_id] = {
                'debt_added': 0,
                'paid': 0,
            }
        if item.get('type') == 'debt':
            agg[supplier_id]['debt_added'] += amount
        elif item.get('type') == 'payment':
            agg[supplier_id]['paid'] += amount

    rows = []
    totals = {'debt_added': 0, 'paid': 0, 'remaining': 0}
    for supplier_id, s in suppliers.items():
        if not isinstance(s, dict):
            continue
        name = str(s.get('name', '—'))
        remaining = int(s.get('balance', 0) or 0)
        debt_added = int(agg.get(supplier_id, {}).get('debt_added', 0) or 0)
        paid = int(agg.get(supplier_id, {}).get('paid', 0) or 0)

        rows.append({
            'supplier_id': supplier_id,
            'supplier_name': name,
            'debt_added': debt_added,
            'paid': paid,
            'remaining': remaining,
        })
        totals['debt_added'] += debt_added
        totals['paid'] += paid
        totals['remaining'] += remaining

    rows.sort(key=lambda r: r['remaining'], reverse=True)
    return rows, totals

# --- KEYBOARDS ---
def get_main_menu(admin: bool = False):
    buttons = [
        [KeyboardButton(text="➕ Добавить поставщика"), KeyboardButton(text="💰 Добавить долг")],
        [KeyboardButton(text="💸 Погасить долг"), KeyboardButton(text="📋 Список поставщиков")],
        [KeyboardButton(text="📊 Отчеты")]
    ]
    if admin:
        buttons.append([KeyboardButton(text="🗑 Очистить базу")])
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)

def get_reports_menu():
    buttons = [
        [KeyboardButton(text="📊 Скачать Excel"), KeyboardButton(text="📄 Скачать PDF")],
        [KeyboardButton(text="⬅️ Назад")]
    ]
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)

# --- BOT LOGIC ---
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()
dp.message.middleware(AuthMiddleware())

@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    user_id = message.from_user.id if message.from_user else 0
    if await is_authorized(user_id):
        await message.answer(
            "🏢 Добро пожаловать в систему учета долгов **Olov**.\n"
            "Выберите действие в меню ниже:",
            reply_markup=get_main_menu(admin=is_admin(user_id)),
            parse_mode="Markdown"
        )
        return

    await message.answer(
        "🏢 Добро пожаловать в систему учета долгов **Olov**.\n"
        "Для продолжения нужен PIN-код доступа.",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="Markdown"
    )

    # переводим в режим ввода PIN
    # (если state еще не создан, middleware поставит его при первом сообщении)


@dp.message(Form.auth_pin)
async def auth_pin_enter(message: types.Message, state: FSMContext):
    pin = (message.text or "").strip()
    if pin != ACCESS_PIN:
        await message.answer("❌ Неверный PIN. Попробуйте еще раз:")
        return

    user_id = message.from_user.id if message.from_user else 0
    await set_authorized(user_id, True)
    await state.clear()
    await message.answer("✅ Доступ предоставлен.", reply_markup=get_main_menu(admin=is_admin(user_id)))

# --- SUPPLIER MANAGEMENT ---
@dp.message(F.func(cmd_filter("➕ Добавить поставщика", "добавить поставщика")))
async def add_supplier_start(message: types.Message, state: FSMContext):
    await state.set_state(Form.add_supplier)
    await message.answer("Введите имя нового поставщика:", reply_markup=ReplyKeyboardRemove())

@dp.message(Form.add_supplier)
async def add_supplier_finish(message: types.Message, state: FSMContext):
    name = message.text.strip()
    if not name:
        await message.answer("Имя не может быть пустым.")
        return

    if firebase is None:
        await message.answer("❌ Firebase не инициализирован.")
        return

    created = await firebase.post("suppliers", {
        'name': name,
        'balance': 0,
        'created_at': datetime.now().isoformat()
    })
    _ = created.get("name")
    
    await state.clear()
    await message.answer(f"✅ Поставщик '{name}' успешно добавлен!", reply_markup=get_main_menu(admin=is_admin(message.from_user.id)))

@dp.message(F.func(cmd_filter("📋 Список поставщиков", "список поставщиков")))
async def list_suppliers(message: types.Message):
    if firebase is None:
        await message.answer("❌ Firebase не инициализирован.")
        return

    suppliers = await firebase.get('suppliers')
    
    if not suppliers:
        await message.answer("Список поставщиков пуст.")
        return

    text = "📋 **Список поставщиков:**\n\n"
    total_debt = 0
    for key, data in suppliers.items():
        balance = data.get('balance', 0)
        total_debt += balance
        text += f"👤 {data['name']}: {format_currency(balance)}\n"
    
    text += f"\n--- \n💰 **Общий долг:** {format_currency(total_debt)}"
    await message.answer(text, parse_mode="Markdown")

# --- DEBT MANAGEMENT ---
@dp.message(F.func(cmd_filter("💰 Добавить долг", "добавить долг")))
async def add_debt_start(message: types.Message, state: FSMContext):
    if firebase is None:
        await message.answer("❌ Firebase не инициализирован.")
        return

    suppliers = await firebase.get('suppliers')
    if not suppliers:
        await message.answer("Сначала добавьте поставщика.")
        return
    
    kb = []
    for key, data in suppliers.items():
        kb.append([KeyboardButton(text=data['name'])])
    kb.append([KeyboardButton(text="❌ Отмена")])
    
    await state.set_state(Form.select_supplier_for_debt)
    await message.answer("Выберите поставщика:", reply_markup=ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True))

@dp.message(Form.select_supplier_for_debt)
async def add_debt_supplier_selected(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=get_main_menu(admin=is_admin(message.from_user.id)))
        return

    await state.update_data(supplier_name=message.text)
    await state.set_state(Form.amount_debt)
    await message.answer(f"Введите сумму долга для '{message.text}':\n(Например: 100 000 или 50.000)")

@dp.message(Form.amount_debt)
async def add_debt_finish(message: types.Message, state: FSMContext):
    amount = clean_amount(message.text)
    if amount <= 0:
        await message.answer("Сумма должна быть больше нуля.")
        return

    if firebase is None:
        await message.answer("❌ Firebase не инициализирован.")
        return

    data = await state.get_data()
    supplier_name = data['supplier_name']
    
    suppliers = await firebase.get('suppliers')
    
    supplier_id = None
    current_balance = 0
    for key, val in suppliers.items():
        if val['name'] == supplier_name:
            supplier_id = key
            current_balance = val.get('balance', 0)
            break
            
    if supplier_id:
        new_balance = current_balance + amount
        await firebase.patch(f"suppliers/{supplier_id}", {'balance': new_balance})
        
        # Log history
        await firebase.post('history', {
            'supplier_id': supplier_id,
            'supplier_name': supplier_name,
            'type': 'debt',
            'amount': amount,
            'date': datetime.now().isoformat()
        })
        
        await state.clear()
        await message.answer(f"✅ Долг {format_currency(amount)} добавлен.\nНовый баланс: {format_currency(new_balance)}", reply_markup=get_main_menu(admin=is_admin(message.from_user.id)))

# --- PAYMENT MANAGEMENT ---
@dp.message(F.func(cmd_filter("💸 Погасить долг", "погасить долг")))
async def pay_debt_start(message: types.Message, state: FSMContext):
    if firebase is None:
        await message.answer("❌ Firebase не инициализирован.")
        return

    suppliers = await firebase.get('suppliers')
    if not suppliers:
        await message.answer("Нет поставщиков.")
        return
    
    kb = []
    for key, data in suppliers.items():
        kb.append([KeyboardButton(text=data['name'])])
    kb.append([KeyboardButton(text="❌ Отмена")])
    
    await state.set_state(Form.select_supplier_for_payment)
    await message.answer("Выберите поставщика для оплаты:", reply_markup=ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True))

@dp.message(Form.select_supplier_for_payment)
async def pay_debt_supplier_selected(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=get_main_menu(admin=is_admin(message.from_user.id)))
        return

    await state.update_data(supplier_name=message.text)
    await state.set_state(Form.amount_payment)
    await message.answer(f"Введите сумму оплаты для '{message.text}':")

@dp.message(Form.amount_payment)
async def pay_debt_finish(message: types.Message, state: FSMContext):
    amount = clean_amount(message.text)
    if amount <= 0:
        await message.answer("Сумма должна быть больше нуля.")
        return

    if firebase is None:
        await message.answer("❌ Firebase не инициализирован.")
        return

    data = await state.get_data()
    supplier_name = data['supplier_name']
    
    suppliers = await firebase.get('suppliers')
    
    supplier_id = None
    current_balance = 0
    for key, val in suppliers.items():
        if val['name'] == supplier_name:
            supplier_id = key
            current_balance = val.get('balance', 0)
            break
            
    if supplier_id:
        new_balance = current_balance - amount
        await firebase.patch(f"suppliers/{supplier_id}", {'balance': new_balance})
        
        # Log history
        await firebase.post('history', {
            'supplier_id': supplier_id,
            'supplier_name': supplier_name,
            'type': 'payment',
            'amount': amount,
            'date': datetime.now().isoformat()
        })
        
        await state.clear()
        await message.answer(f"✅ Оплата {format_currency(amount)} зафиксирована.\nОстаток долга: {format_currency(new_balance)}", reply_markup=get_main_menu(admin=is_admin(message.from_user.id)))

# --- REPORTS ---
@dp.message(F.func(cmd_filter("📊 Отчеты", "отчеты", "отчёты")))
async def reports_menu(message: types.Message):
    await message.answer("Выберите формат отчета:", reply_markup=get_reports_menu())

@dp.message(F.func(cmd_filter("⬅️ Назад", "назад")))
async def back_to_main(message: types.Message):
    await message.answer("Главное меню:", reply_markup=get_main_menu(admin=is_admin(message.from_user.id)))


@dp.message(F.func(cmd_filter("� Очистить базу", "очистить базу")))
async def admin_wipe_start(message: types.Message, state: FSMContext):
    user_id = message.from_user.id if message.from_user else 0
    if not is_admin(user_id):
        await message.answer("❌ Недостаточно прав.")
        return
    await state.set_state(Form.admin_wipe_pin)
    await message.answer("⚠️ Введите PIN для удаления базы:", reply_markup=ReplyKeyboardRemove())


@dp.message(Form.admin_wipe_pin)
async def admin_wipe_confirm(message: types.Message, state: FSMContext):
    user_id = message.from_user.id if message.from_user else 0
    if not is_admin(user_id):
        await state.clear()
        await message.answer("❌ Недостаточно прав.", reply_markup=get_main_menu(admin=False))
        return

    pin = (message.text or "").strip()
    if pin != WIPE_PIN:
        await message.answer("❌ Неверный PIN. Попробуйте еще раз:")
        return

    if firebase is None:
        await state.clear()
        await message.answer("❌ Firebase не инициализирован.", reply_markup=get_main_menu(admin=True))
        return

    try:
        # Очищаем основные ветки данных
        await firebase.delete("suppliers")
        await firebase.delete("history")

        # Сбрасываем кэш авторизации, но админ останется авторизованным в Firebase
        _auth_cache.clear()
        await state.clear()
        await message.answer("✅ База очищена (suppliers, history).", reply_markup=get_main_menu(admin=True))
    except Exception as e:
        logger.error(f"Admin wipe failed: {e}")
        await state.clear()
        await message.answer(f"❌ Ошибка очистки базы: {e}", reply_markup=get_main_menu(admin=True))

@dp.message(F.func(cmd_filter("�📊 Скачать Excel", "скачать excel", "excel")))
async def export_excel(message: types.Message):
    if firebase is None:
        await message.answer("❌ Firebase не инициализирован.")
        return

    rows, totals = await build_detailed_report()
    if not rows:
        await message.answer("Данных для экспорта нет.")
        return
    
    data_list = []
    for r in rows:
        data_list.append({
            "Поставщик": r['supplier_name'],
            "Начислено (долг), UZS": r['debt_added'],
            "Оплачено, UZS": r['paid'],
            "Остаток к оплате, UZS": r['remaining'],
        })
    
    df = pd.DataFrame(data_list)
    filename = f"reports/report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    os.makedirs('reports', exist_ok=True)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Отчет")
        ws = writer.book["Отчет"]

        totals_row = ws.max_row + 2
        ws.cell(row=totals_row, column=1, value="ИТОГО").font = Font(bold=True)
        ws.cell(row=totals_row, column=2, value=totals['debt_added']).font = Font(bold=True)
        ws.cell(row=totals_row, column=3, value=totals['paid']).font = Font(bold=True)
        ws.cell(row=totals_row, column=4, value=totals['remaining']).font = Font(bold=True)

        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                c.border = border
                if row > 1 and col >= 2:
                    c.number_format = '#,##0'
                    c.alignment = Alignment(horizontal="right")
                elif row > 1 and col == 1:
                    c.alignment = Alignment(horizontal="left")

        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 22
    
    await message.answer_document(FSInputFile(filename), caption="📊 Отчет Excel готов.")

@dp.message(F.func(cmd_filter("📄 Скачать PDF", "скачать pdf", "pdf")))
async def export_pdf(message: types.Message):
    try:
        if firebase is None:
            await message.answer("❌ Firebase не инициализирован.")
            return

        rows, totals = await build_detailed_report()
        if not rows:
            await message.answer("Данных для экспорта нет.")
            return
        
        report_dt = datetime.now()
        filename = f"reports/report_{report_dt.strftime('%Y%m%d_%H%M')}.pdf"
        os.makedirs('reports', exist_ok=True)

        doc = SimpleDocTemplate(
            filename,
            pagesize=A4,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
            title="Olov - Debt Report",
            author="Olov"
        )
        elements = []
        styles = getSampleStyleSheet()

        # Стиль: официальный, минимализм
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Normal'],
            fontName=DEFAULT_FONT,
            fontSize=16,
            leading=20,
            alignment=0,
            spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontName=DEFAULT_FONT,
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#4B5563"),
            spaceAfter=14
        )
        body_style = ParagraphStyle(
            'ReportBody',
            parent=styles['Normal'],
            fontName=DEFAULT_FONT,
            fontSize=10,
            leading=13
        )
        total_style = ParagraphStyle(
            'ReportTotal',
            parent=styles['Normal'],
            fontName=DEFAULT_FONT,
            fontSize=12,
            leading=16,
            spaceBefore=12
        )

        elements.append(Paragraph("Olov", title_style))
        elements.append(Paragraph("Официальный отчет по задолженностям поставщиков", subtitle_style))
        elements.append(Paragraph(f"Дата формирования: {report_dt.strftime('%d.%m.%Y %H:%M')}", subtitle_style))

        # Таблица
        table_data = [["Поставщик", "Начислено, UZS", "Оплачено, UZS", "Остаток, UZS"]]
        for r in rows:
            table_data.append([
                r['supplier_name'],
                format_currency(r['debt_added']),
                format_currency(r['paid']),
                format_currency(r['remaining']),
            ])

        t = Table(table_data, colWidths=[7.6 * cm, 3.1 * cm, 3.1 * cm, 3.1 * cm], hAlign='LEFT')

        header_bg = colors.HexColor("#111827")
        stripe_bg = colors.HexColor("#F3F4F6")
        grid_color = colors.HexColor("#D1D5DB")

        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), DEFAULT_FONT),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), DEFAULT_FONT),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('LINEBELOW', (0, 0), (-1, 0), 0.8, colors.HexColor("#374151")),
            ('GRID', (0, 0), (-1, -1), 0.5, grid_color),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, stripe_bg]),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ]
        t.setStyle(TableStyle(table_style))

        elements.append(t)
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"Начислено всего: <b>{format_currency(totals['debt_added'])}</b>", body_style))
        elements.append(Paragraph(f"Оплачено всего: <b>{format_currency(totals['paid'])}</b>", body_style))
        elements.append(Paragraph(f"Остаток к оплате: <b>{format_currency(totals['remaining'])}</b>", total_style))

        def _on_page(canv: canvas.Canvas, doc_):
            canv.saveState()
            canv.setFont(DEFAULT_FONT, 9)
            canv.setFillColor(colors.HexColor("#6B7280"))
            footer_left = "Olov • Отчет по задолженностям"
            footer_right = f"Стр. {canv.getPageNumber()}"
            canv.drawString(doc_.leftMargin, 1.2 * cm, footer_left)
            canv.drawRightString(doc_.pagesize[0] - doc_.rightMargin, 1.2 * cm, footer_right)
            canv.restoreState()

        doc.build(elements, onFirstPage=_on_page, onLaterPages=_on_page)
        
        await message.answer_document(FSInputFile(filename), caption="📄 Отчет PDF готов.")
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        await message.answer(f"❌ Ошибка при генерации PDF: {e}")

# --- MAIN ---
async def main():
    logger.info("Starting bot...")
    global firebase, firebase_session
    firebase_session = aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=20))
    firebase = FirebaseREST(FIREBASE_URL, firebase_session)
    try:
        await dp.start_polling(bot)
    finally:
        if firebase_session is not None:
            await firebase_session.close()


@dp.message()
async def unknown_message(message: types.Message):
    await message.answer("Не понял команду. Используй кнопки меню ниже.", reply_markup=get_main_menu(admin=is_admin(message.from_user.id)))

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logger.info("Bot stopped.")


